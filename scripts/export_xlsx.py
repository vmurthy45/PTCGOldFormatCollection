"""Export the whole collection to a single .xlsx workbook.

Since the site's JSON files are now the source of truth (maintained from chat),
this is how Vig gets a spreadsheet copy back out — a full snapshot he can keep,
sort, or hand to anything that wants Excel.

    .venv/bin/python scripts/export_xlsx.py [output.xlsx]

Sheets:
    Inventory     one row per (card, print/version) he owns, with which decks
                  use that card, or "Spare" when no deck does
    Decks         every card line of every deck, with print-type breakdown
    Cards to Get  missing copies aggregated across decks (the shopping list)
    Tournaments   the tournament log
    Summary       headline counts
"""
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DEFAULT_OUT = ROOT / f"PTCG Collection Export {date.today():%Y-%m-%d}.xlsx"


def load(name, default):
    p = DATA / name
    if not p.exists():
        return default
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return default


def name_key(s):
    s = unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode()
    s = s.lower().replace("’", "'")
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"^basic\s+", "", s)
    return re.sub(r"[^a-z0-9]+", "", s)


def missing_of(c):
    p = c.get("prints") or {}
    owned = (p.get("authentic", 0) or 0) + (p.get("wc", 0) or 0) + (p.get("proxy", 0) or 0)
    return max(0, int(c.get("count", 0)) - owned)


def main():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: .venv/bin/pip install openpyxl")

    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUT
    cards = load("cards.json", [])
    inventory = load("inventory.json", [])
    tournaments = load("tournaments.json", [])

    # card name -> {deck: copies}, for the inventory sheet's "used in" column
    deck_use = defaultdict(lambda: defaultdict(int))
    for c in cards:
        deck_use[name_key(c["name"])][c["deck"]] += int(c.get("count", 0))

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="EE2B39")

    def sheet(title, headers, rows, widths):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    # --- Inventory ---
    inv_rows = []
    for g in inventory:
        decks = deck_use.get(g.get("key", name_key(g["name"])), {})
        used = "; ".join(f"{d} x{n}" for d, n in
                         sorted(decks.items(), key=lambda kv: (-kv[1], kv[0]))) or "Spare"
        for v in g.get("variants", []):
            inv_rows.append([
                g["name"], g.get("type", ""), v.get("version", ""),
                v.get("set", ""), v.get("num", ""), v.get("qty", 0),
                len(decks), used,
            ])
    sheet("Inventory",
          ["Card", "Type", "Version", "Set", "Number", "Qty", "Decks", "Used In"],
          inv_rows, [30, 12, 15, 8, 9, 6, 7, 80])

    # --- Decks ---
    deck_rows = []
    for c in cards:
        p = c.get("prints") or {}
        deck_rows.append([
            c.get("year", ""), c.get("deck", ""), c.get("count", 0), c.get("name", ""),
            c.get("set", ""), c.get("category", ""),
            p.get("authentic", 0), p.get("wc", 0), p.get("proxy", 0), missing_of(c),
        ])
    sheet("Decks",
          ["Format", "Deck", "Count", "Card", "Set", "Category",
           "Standard", "WC", "Proxy", "Missing"],
          deck_rows, [9, 26, 7, 28, 11, 11, 10, 7, 8, 9])

    # --- Cards to Get ---
    need = defaultdict(lambda: {"total": 0, "where": []})
    for c in cards:
        m = missing_of(c)
        if m:
            e = need[c["name"]]
            e["total"] += m
            e["where"].append(f"{c['deck']} x{m}")
    get_rows = [[n, v["total"], "; ".join(v["where"])]
                for n, v in sorted(need.items(), key=lambda kv: (-kv[1]["total"], kv[0]))]
    sheet("Cards to Get", ["Card", "Copies Needed", "Needed For"],
          get_rows, [30, 14, 80])

    # --- Tournaments ---
    t_rows = []
    for t in tournaments:
        rec = t.get("record") or {}
        t_rows.append([
            t.get("date", ""), t.get("name", ""), t.get("deck", ""),
            t.get("players", ""), rec.get("wins", ""), rec.get("losses", ""), rec.get("ties", ""),
            t.get("placement", ""), t.get("entryCost", ""), t.get("prizesWon", ""),
            t.get("notes", ""),
        ])
    sheet("Tournaments",
          ["Date", "Event", "Deck", "Players", "W", "L", "T", "Placement",
           "Entry Cost", "Prizes Won", "Notes"],
          t_rows, [12, 40, 24, 9, 5, 5, 5, 11, 11, 12, 40])

    # --- Summary ---
    decks = {c["deck"] for c in cards}
    total_copies = sum(int(c.get("count", 0)) for c in cards)
    total_missing = sum(missing_of(c) for c in cards)
    owned_copies = sum(v.get("qty", 0) for g in inventory for v in g.get("variants", []))
    spare = sum(1 for g in inventory if not deck_use.get(g.get("key", name_key(g["name"]))))
    sheet("Summary", ["Metric", "Value"], [
        ["Exported", f"{date.today():%Y-%m-%d}"],
        ["Decks", len(decks)],
        ["Deck card lines", len(cards)],
        ["Cards across all decks", total_copies],
        ["Copies still missing", total_missing],
        ["Inventory cards (unique)", len(inventory)],
        ["Inventory copies owned", owned_copies],
        ["Inventory cards not in any deck (spare)", spare],
        ["Tournaments logged", len(tournaments)],
    ], [40, 16])

    wb.remove(wb["Sheet"])
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  Inventory {len(inv_rows)} rows · Decks {len(deck_rows)} rows · "
          f"Cards to Get {len(get_rows)} rows · Tournaments {len(t_rows)} rows")


if __name__ == "__main__":
    main()
